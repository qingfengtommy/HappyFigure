#!/usr/bin/env python3
"""
Lightweight data inspector for figure planning.

Usage:
    python inspect_data.py <directory>

Scans a directory recursively and prints a structured summary of every data file,
including column names, row counts, sheet names, and inferred plot-type suggestions.
"""

import sys
import os
import json
from pathlib import Path

def inspect_excel(path):
    """Inspect an Excel workbook."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            rows = list(ws.iter_rows(max_row=6, values_only=True))
            if not rows:
                sheets.append({"sheet": name, "empty": True})
                continue
            # First non-None row as headers
            headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]
            nrows = ws.max_row - 1 if ws.max_row else 0
            ncols = ws.max_column or 0
            # Also capture first few row labels (column A values after header)
            # This helps detect transposed paired data where conditions are in rows
            row_labels = []
            for r in rows[1:]:
                if r and r[0] is not None:
                    row_labels.append(str(r[0]).strip())
            info = {
                "sheet": name,
                "columns": headers[:15],  # cap at 15
                "n_columns_total": ncols,
                "n_rows": nrows,
                "plot_hint": suggest_plot_type(headers, nrows, ncols, row_labels=row_labels)
            }
            if row_labels:
                info["row_labels_sample"] = row_labels[:5]
            sheets.append(info)
        wb.close()
        return {"type": "excel", "sheets": sheets}
    except Exception as e:
        return {"type": "excel", "error": str(e)}


def inspect_csv(path):
    """Inspect a CSV file."""
    try:
        import csv
        with open(path, newline='', encoding='utf-8', errors='replace') as f:
            reader = csv.reader(f)
            headers = next(reader, [])
            nrows = sum(1 for _ in reader)
        return {
            "type": "csv",
            "columns": headers[:15],
            "n_columns": len(headers),
            "n_rows": nrows,
            "plot_hint": suggest_plot_type(headers, nrows, len(headers))
        }
    except Exception as e:
        return {"type": "csv", "error": str(e)}


def inspect_numpy(path):
    """Inspect a numpy file."""
    try:
        import numpy as np
        data = np.load(path, allow_pickle=True)
        if isinstance(data, np.ndarray):
            return {"type": "numpy", "shape": list(data.shape), "dtype": str(data.dtype)}
        else:  # npz
            return {"type": "npz", "arrays": {k: {"shape": list(data[k].shape), "dtype": str(data[k].dtype)} for k in data.files}}
    except Exception as e:
        return {"type": "numpy", "error": str(e)}


def suggest_plot_type(headers, nrows, ncols, row_labels=None):
    """Suggest a plot type based on column names, data shape, and row labels."""
    headers_lower = [h.lower() for h in headers if h]
    row_labels_lower = [r.lower() for r in (row_labels or [])]

    # Check for paired structure in COLUMNS — require explicit before/after naming
    paired_col_keywords = ['before', 'after', 'pre', 'post', 'acute', 'convalescent']
    paired_count = sum(1 for h in headers_lower if any(k in h for k in paired_col_keywords))
    if paired_count >= 2 and nrows < 50:
        return "paired-line plot (matched before/after)"

    # Check for paired structure in ROWS (transposed layout)
    # Pattern: exactly 2-3 data rows, many columns (subjects), row labels are conditions
    if row_labels_lower:
        row_paired = sum(1 for r in row_labels_lower if any(k in r for k in ['baseline', 'before', 'after', 'pre', 'post']))
        if row_paired >= 1 and nrows <= 4 and ncols >= 5:
            return "MAYBE paired-line (transposed: conditions in rows, subjects in columns — verify with text)"

    # Check for time series
    time_keywords = ['trial', 'epoch', 'step', 'day', 'time', 'iteration']
    if any(any(k in h for k in time_keywords) for h in headers_lower):
        return "line plot with error bands (time series)"

    # Check for ordination
    ordination_keywords = ['pc1', 'pc2', 'umap1', 'umap2', 'tsne1', 'tsne2', 'pca']
    if any(any(k in h for k in ordination_keywords) for h in headers_lower):
        return "scatter plot (ordination/embedding)"

    # Check for differential/volcano
    de_keywords = ['log2fc', 'fold_change', 'logfc', 'pvalue', 'p_value', 'padj', 'fdr']
    if sum(1 for h in headers_lower if any(k in h for k in de_keywords)) >= 2:
        return "volcano plot (differential analysis)"

    # Matrix-like: many columns, many rows → heatmap
    if ncols > 8 and nrows > 5:
        return "heatmap (matrix with many columns)"

    # Large n → distribution
    if nrows > 100 and ncols <= 4:
        return "violin plot (large-n distribution)"

    # Medium n, few groups
    if nrows > 30 and ncols <= 6:
        return "violin or box+strip (moderate-n distribution)"

    # Few groups, small n
    if ncols <= 6 and nrows <= 30:
        return "bar chart with individual data points"

    return "bar chart or scatter (inspect data values to refine)"


def scan_directory(root):
    """Scan directory and inspect all data files."""
    results = []
    root = Path(root)

    for path in sorted(root.rglob("*")):
        if not path.is_file():
            continue
        rel = str(path.relative_to(root))
        ext = path.suffix.lower()

        if ext in ('.xlsx', '.xls'):
            info = inspect_excel(str(path))
            info["path"] = rel
            results.append(info)
        elif ext in ('.csv', '.tsv'):
            info = inspect_csv(str(path))
            info["path"] = rel
            results.append(info)
        elif ext in ('.npy', '.npz'):
            info = inspect_numpy(str(path))
            info["path"] = rel
            results.append(info)
        elif ext in ('.json', '.yaml', '.yml'):
            results.append({"type": "config", "path": rel, "size": path.stat().st_size})
        elif ext in ('.md', '.txt', '.tex'):
            results.append({"type": "text", "path": rel, "size": path.stat().st_size})
        elif ext in ('.py', '.r', '.R', '.jl', '.m'):
            results.append({"type": "code", "path": rel, "size": path.stat().st_size})
        elif ext in ('.png', '.jpg', '.jpeg', '.svg', '.tif', '.tiff'):
            results.append({"type": "image", "path": rel, "size": path.stat().st_size})

    return results


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python inspect_data.py <directory>")
        sys.exit(1)

    directory = sys.argv[1]
    if not os.path.isdir(directory):
        print(f"Error: {directory} is not a directory")
        sys.exit(1)

    results = scan_directory(directory)
    print(json.dumps(results, indent=2, default=str))
